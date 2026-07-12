import csv, glob, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(__file__)
OUT_DIR = os.path.join(BASE, "output", "us_local")
EVAL_DIR = os.path.join(BASE, "output", "evaluation")
os.makedirs(EVAL_DIR, exist_ok=True)

# ── GICS sector labels per ETF ──────────────────────────────────────────────
ETF_SECTOR = {
    "XLC": "Communication Services", "XLY": "Consumer Discretionary",
    "XLP": "Consumer Staples",        "XLE": "Energy",
    "XLF": "Financials",              "XLV": "Health Care",
    "XLI": "Industrials",             "XLK": "Technology",
    "XLB": "Materials",               "XLRE": "Real Estate",
    "XLU": "Utilities",               "SOXX": "Semiconductors",
}

def etf_rating(score):
    if score is None: return "N/A"
    if score >= 0.80: return "Very Strong"
    if score >= 0.60: return "Strong"
    if score >= 0.40: return "Neutral"
    if score >= 0.20: return "Weak"
    return "Very Weak"

# ── Load scores_detail files for ETF scores & sector lookup ─────────────────
detail_cache = {}
for strategy in ("momentum", "reversal"):
    for dt in ("2026-05-04", "2026-05-18"):
        path = os.path.join(OUT_DIR, f"scores_detail_{strategy}_{dt}.json")
        if os.path.exists(path):
            with open(path) as fh:
                detail_cache[(strategy, dt)] = json.load(fh)

# ── Read ticker → ETF map from constituents file ─────────────────────────────
ticker_etf = {}
for cand in ("data/constituents_us_combined.csv",
             "pipeline/data/constituents_us_combined.csv",
             "constituents_us_combined.csv"):
    p = os.path.join(BASE, cand)
    if os.path.exists(p):
        with open(p) as fh:
            for row in csv.DictReader(fh):
                etf = row.get("ETF") or row.get("Sector_ETF") or ""
                sym = row.get("Symbol") or row.get("Ticker") or ""
                if sym and etf:
                    ticker_etf[sym.strip()] = etf.strip()
        break

# ── Collect all cap-split rows ───────────────────────────────────────────────
rows = []
CAPS = ("large", "mid", "small")
SIDES = ("bull", "bear")
STRATEGIES = ("momentum", "reversal")
ML_TYPES = ("composite", "pureml")

for dt in ("2026-05-04", "2026-05-18"):
    for strategy in STRATEGIES:
        detail = detail_cache.get((strategy, dt), {})
        for ml in ML_TYPES:
            for side in SIDES:
                for cap in CAPS:
                    fname = f"watchlist_{strategy}_{ml}_{side}_{cap}_{dt}.csv"
                    fpath = os.path.join(OUT_DIR, fname)
                    if not os.path.exists(fpath):
                        continue
                    with open(fpath) as fh:
                        reader = csv.DictReader(fh)
                        for r in reader:
                            ticker = r.get("ticker", "").strip()
                            if not ticker:
                                continue
                            d_side = side  # bull or bear
                            tk_data = detail.get(ticker, {}).get(d_side, {})
                            sv = tk_data.get("signal_values", {})
                            etf = ticker_etf.get(ticker, "")
                            etf_score_raw = sv.get("sector_etf_bull_score")
                            rows.append({
                                "Date": dt,
                                "Strategy": strategy.capitalize(),
                                "ML Type": "Pure ML" if ml == "pureml" else "Composite",
                                "Side": side.capitalize(),
                                "Cap Tier": cap.capitalize() + " Cap",
                                "Rank": int(r.get("rank", 0)),
                                "Ticker": ticker,
                                "Score": float(r.get("score", 0)),
                                "Model Score": round(tk_data.get("model_score", 0), 4),
                                "Sector ETF": etf,
                                "Sector": ETF_SECTOR.get(etf, etf or "Unknown"),
                                "Sector ETF Score": round(etf_score_raw, 4) if etf_score_raw is not None else None,
                                "Sector Rating": etf_rating(etf_score_raw),
                                "SDZ HTF": float(r.get("sdz_htf_score", 0)),
                                "SSZ HTF": float(r.get("ssz_htf_score", 0)),
                                "Zone Confluence": float(r.get("zone_htf_confluence", 0)),
                                "Sector RS 20d": float(r.get("sector_rs_20d", 0)),
                                "Return 20d": float(r.get("return_20d", 0)),
                                "ADX 14": float(r.get("adx_14", 0)),
                                "Vol Contraction": float(r.get("vol_contraction", 0)),
                            })

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "May 2026 Watchlist"

HEADERS = [
    "Date", "Strategy", "ML Type", "Side", "Cap Tier", "Rank",
    "Ticker", "Score", "Model Score",
    "Sector ETF", "Sector", "Sector ETF Score", "Sector Rating",
    "SDZ HTF", "SSZ HTF", "Zone Confluence",
    "Sector RS 20d", "Return 20d", "ADX 14", "Vol Contraction",
]

# colour palette
C_HEADER_BG  = "1F3864"   # dark navy
C_HEADER_FG  = "FFFFFF"
C_BULL_BG    = "E2EFDA"   # light green
C_BEAR_BG    = "FCE4D6"   # light salmon
C_ALT_BULL   = "F0F7EC"
C_ALT_BEAR   = "FDF1EA"
C_ACCENT     = "2E75B6"   # blue for score cells
C_GOLD       = "FFD966"   # top-3 rank highlight

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Header row
for col_idx, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=hdr)
    cell.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=10)
    cell.fill = PatternFill("solid", start_color=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

ws.row_dimensions[1].height = 30

# Data rows
for row_i, r in enumerate(rows, 2):
    is_bull = r["Side"] == "Bull"
    is_alt  = (row_i % 2 == 0)
    if is_bull:
        bg = C_BULL_BG if not is_alt else C_ALT_BULL
    else:
        bg = C_BEAR_BG if not is_alt else C_ALT_BEAR
    fill = PatternFill("solid", start_color=bg)

    for col_idx, hdr in enumerate(HEADERS, 1):
        val = r[hdr]
        cell = ws.cell(row=row_i, column=col_idx, value=val)
        cell.font = Font(name="Arial", size=9)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # rank top-3 gold highlight
        if hdr == "Rank" and isinstance(val, int) and val <= 3:
            cell.fill = PatternFill("solid", start_color=C_GOLD)
            cell.font = Font(name="Arial", size=9, bold=True)

        # score columns → 4 decimals
        if hdr in ("Score", "Model Score", "Sector ETF Score"):
            cell.number_format = "0.0000"
        elif hdr in ("SDZ HTF", "SSZ HTF", "Zone Confluence"):
            cell.number_format = "0.00"
        elif hdr in ("Sector RS 20d", "Return 20d"):
            cell.number_format = "0.0000"
        elif hdr in ("ADX 14", "Vol Contraction"):
            cell.number_format = "0.00"

# Column widths
widths = {
    "Date": 12, "Strategy": 12, "ML Type": 12, "Side": 8, "Cap Tier": 11,
    "Rank": 6, "Ticker": 8, "Score": 9, "Model Score": 11,
    "Sector ETF": 10, "Sector": 22, "Sector ETF Score": 14, "Sector Rating": 13,
    "SDZ HTF": 9, "SSZ HTF": 9, "Zone Confluence": 13,
    "Sector RS 20d": 12, "Return 20d": 11, "ADX 14": 9, "Vol Contraction": 13,
}
for col_idx, hdr in enumerate(HEADERS, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(hdr, 12)

# Auto-filter
ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = "A2"

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
summary_hdrs = ["Date", "Strategy", "ML Type", "Side", "Cap Tier", "Count"]
for ci, h in enumerate(summary_hdrs, 1):
    c = ws2.cell(row=1, column=ci, value=h)
    c.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=10)
    c.fill = PatternFill("solid", start_color=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border

from collections import Counter
cnt = Counter((r["Date"], r["Strategy"], r["ML Type"], r["Side"], r["Cap Tier"]) for r in rows)
for ri, (key, n) in enumerate(sorted(cnt.items()), 2):
    for ci, v in enumerate(list(key) + [n], 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.font = Font(name="Arial", size=9)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

for ci, h in enumerate(summary_hdrs, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = 14

out_path = os.path.join(EVAL_DIR, "may_2026_watchlist_full.xlsx")
wb.save(out_path)
print(f"Saved {len(rows)} rows -> {out_path}")
